"""
Phase A 표본 테스트 — 5건으로 새 카드 구조 검증

목적:
1. 시트별 컬럼 매핑 작동 확인 (1·2·3·4차 모두)
2. rich text 색 정보 추출 (v2 파랑·v3 빨강 마커)
3. 사용자 통찰 직접 검증 — 협의결과=blank·TOBE 채워진 행도 학습 자료
4. 시트 4 customer_confirm·additional_suggest 메타 추가
5. NEGATIVE 분류 정확성 (기존유지·미수용·원안유지)

표본:
- 1차 LN102 (NEGATIVE 검증)
- 1차 (협의=blank, TOBE 채워짐) 1건 — 사용자 통찰 직접 증거
- 2차 AC283 (MEDIUM 검증)
- 3차 첫 행 (msg_code 없음, msg_name 기반)
- 4차 AP007 (customer_confirm·additional_suggest 메타)

출력: output/_표본/ 폴더에 5개 카드 .txt (사용자 검토용)
"""

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from collections import defaultdict

XLSX_PATH = Path(r"C:\dev\guide\정답데이터2\정답데이터2.xlsx")
OUTPUT_DIR = Path(r"C:\dev\guide\output\_표본")

# 시트별 컬럼 매핑 (1행 헤더 정확 분석 결과)
SHEET_CFG = {
    "1차_국내해외주식채권_피드백반영": {
        "msg_code": 2, "msg_name": 3, "asis": 8, "tobe": 9,
        "feedback": 10, "team": 16, "review": 18, "agree": 21,
    },
    "2차_파생 등_피드백반영": {
        "msg_code": 2, "msg_name": 3, "asis": 8, "tobe": 9,
        "team": 14, "review": 16, "agree": 18, "feedback": 19,
        "wirelink_memo": 20,  # 2차만 별도
    },
    "3차_프로세스 외_피드백 반영": {
        "msg_name": 2, "asis": 3, "tobe": 4, "feedback": 5,
        "team": 6, "agree": 8, "memo": 9,
        # msg_code 없음 — 3차는 msg_name으로 식별
    },
    "4차_ 연금 전체": {
        "msg_code": 2, "msg_name": 3, "asis": 8, "tobe": 9,
        "customer_confirm": 10, "additional_suggest": 11,
        "team": 12, "review": 14, "agree": 16, "feedback": 17,
    },
}

NEGATIVE_AGREE = {"기존유지", "삭제"}
NEGATIVE_FEEDBACK_KW = {"미수용", "원안유지"}
POSITIVE_FEEDBACK_KW = {"수용", "이견없음", "이상없음", "이의없음", "진행요청"}


# ============== rich text 색 정보 추출 (xlsx ZIP 직접 파싱) ==============
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def normalize_color(hex_or_indexed):
    """색 문자열 정규화"""
    if not hex_or_indexed:
        return "DEFAULT"
    s = str(hex_or_indexed).upper()
    # ARGB → RGB
    if len(s) == 8:
        s = s[2:]
    if s in ("FF0000", "C00000", "FF2600", "990000", "B30000"):
        return "RED"
    if s in ("0070C0", "0066CC", "0033CC", "0000FF", "1F4E79", "2E75B6"):
        return "BLUE"
    if s in ("000000", "DEFAULT", "1F1F1F"):
        return "BLACK"
    return s


def parse_rich_text_segments(rt_element):
    """sharedStrings.xml의 rich text element를 segments로 분해"""
    segments = []
    for r in rt_element.findall("main:r", NS):
        text = ""
        t = r.find("main:t", NS)
        if t is not None and t.text:
            text = t.text

        color = "DEFAULT"
        rpr = r.find("main:rPr", NS)
        if rpr is not None:
            color_elem = rpr.find("main:color", NS)
            if color_elem is not None:
                rgb = color_elem.get("rgb")
                if rgb:
                    color = normalize_color(rgb)

        if text:
            segments.append({"text": text, "color": color})

    # rich text 없고 단순 t만 있는 경우
    if not segments:
        t = rt_element.find("main:t", NS)
        if t is not None and t.text:
            segments.append({"text": t.text, "color": "DEFAULT"})

    return segments


def load_shared_strings_segments(xlsx_path):
    """xlsx의 sharedStrings.xml을 파싱해 idx → segments 맵 반환"""
    with zipfile.ZipFile(xlsx_path, "r") as z:
        try:
            with z.open("xl/sharedStrings.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
        except KeyError:
            return {}

    result = {}
    for idx, si in enumerate(root.findall("main:si", NS)):
        segments = parse_rich_text_segments(si)
        result[idx] = segments
    return result


def get_cell_string_index(xlsx_path, sheet_name, row, col):
    """특정 셀의 sharedString 인덱스 반환 (없으면 None)"""
    safe_name = sheet_name
    sheet_xml_paths = []

    with zipfile.ZipFile(xlsx_path, "r") as z:
        # workbook.xml에서 시트 ID 매핑
        with z.open("xl/workbook.xml") as f:
            wb_root = ET.parse(f).getroot()
        sheet_id_map = {}
        for s in wb_root.findall("main:sheets/main:sheet", NS):
            name = s.get("name")
            sid = s.get("sheetId")
            r_id = s.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            sheet_id_map[name] = (sid, r_id)

        # rels에서 sheet 파일 경로 매핑
        with z.open("xl/_rels/workbook.xml.rels") as f:
            rels_root = ET.parse(f).getroot()
        rid_to_target = {}
        rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        for r in rels_root.findall(f"{{{rels_ns}}}Relationship"):
            rid_to_target[r.get("Id")] = r.get("Target")

        if safe_name not in sheet_id_map:
            return None
        _, r_id = sheet_id_map[safe_name]
        target = rid_to_target.get(r_id)
        if not target:
            return None

        sheet_path = f"xl/{target}" if not target.startswith("xl/") else target
        with z.open(sheet_path) as f:
            sheet_root = ET.parse(f).getroot()

        col_letter = chr(ord("A") + col - 1) if col <= 26 else None
        if not col_letter:
            return None
        cell_ref = f"{col_letter}{row}"

        for c in sheet_root.findall("main:sheetData/main:row/main:c", NS):
            if c.get("r") == cell_ref and c.get("t") == "s":
                v = c.find("main:v", NS)
                if v is not None and v.text:
                    return int(v.text)
        return None


def extract_cell_segments(xlsx_path, shared_strings, sheet_name, row, col):
    """특정 셀의 segments 추출 (rich text 색 정보 포함)"""
    idx = get_cell_string_index(xlsx_path, sheet_name, row, col)
    if idx is None:
        # 단순 inline string 또는 빈 셀
        return []
    return shared_strings.get(idx, [])


def render_inline_marked(segments):
    """segments → 인라인 마커 텍스트"""
    if not segments:
        return ""
    parts = []
    for s in segments:
        text = s["text"]
        color = s["color"]
        if color == "BLUE":
            parts.append(f"<v2>{text}</v2>")
        elif color == "RED":
            parts.append(f"<v3>{text}</v3>")
        else:
            parts.append(text)
    return "".join(parts)


# ============== 라벨 분류 (사용자 통찰 적용) ==============
def is_negative(agree, feedback):
    """NEGATIVE 분류 — 사용자 통찰: 명시된 원안유지/미수용만"""
    a = str(agree or "").strip()
    f = str(feedback or "").strip()
    if a in NEGATIVE_AGREE:
        return True
    if any(kw in a for kw in ["기존유지", "삭제"]):
        return True
    if any(kw in f for kw in NEGATIVE_FEEDBACK_KW):
        return True
    return False


def classify_label(agree, feedback, review):
    """카드 라벨 — EVAL_ONLY 폐지, NEGATIVE만 분리, 나머지는 모두 카드화"""
    if is_negative(agree, feedback):
        return "NEGATIVE"
    a = str(agree or "").strip()
    r = str(review or "").strip()
    f = str(feedback or "").strip()

    # HIGH: 명시된 합의
    if a == "반영" and any(kw in f for kw in POSITIVE_FEEDBACK_KW):
        return "HIGH"
    if a == "반영":
        return "HIGH"

    # MEDIUM: 검토완료
    if r in ("검토완료", "검수완료"):
        return "MEDIUM"
    if any(kw in f for kw in POSITIVE_FEEDBACK_KW):
        return "MEDIUM"

    # LOW: 검토필요·blank·기타 (이전 EVAL_ONLY 흡수)
    return "LOW"


# ============== 도메인 분류 ==============
TEAM_DOMAIN = {
    "증권운영팀": "credit_loan", "신용관리팀": "credit_loan",
    "파생상품팀": "derivatives", "파생서비스팀": "derivatives",
    "국제거래팀": "derivatives", "CFD팀": "derivatives",
    "랩어카운트팀": "product", "상품서비스팀": "product",
    "신상품팀": "product", "투자상품팀": "product",
    "결제팀": "settlement", "결제서비스팀": "settlement",
    "연금업무개발팀": "pension", "퇴직연금팀": "pension",
    "연금사업본부": "pension", "연금영업팀": "pension",
    "디지털마케팅팀": "marketing", "브랜드마케팅팀": "marketing",
    "채널서비스팀": "process", "디지털기획팀": "process",
    "개인정보보호팀": "process", "고객만족팀": "process",
}


def classify_domain(team, msg_name, sheet_name):
    """도메인 분류 (담당팀 + 메시지 이름 기반)"""
    if "3차_프로세스" in sheet_name:
        return "process"  # 시트 3은 모두 process
    if "4차_ 연금" in sheet_name:
        return "pension"

    team_str = str(team or "").strip()
    msg_str = str(msg_name or "").strip()

    # 메시지 이름 키워드 우선
    name_lower = msg_str.lower()
    if any(kw in msg_str for kw in ["연금", "IRP", "퇴직"]) or "irp" in name_lower:
        return "pension"
    if any(kw in msg_str for kw in ["이벤트", "마케팅", "광고", "혜택"]):
        return "marketing"

    # 담당팀 기반
    for key, dom in TEAM_DOMAIN.items():
        if key in team_str:
            return dom
    return "misc"


# ============== 카드 본문 생성 ==============
def make_card_text(card):
    """카드 한 건의 본문 텍스트 생성"""
    lines = []
    lines.append("━" * 67)
    lines.append(f"### 사례 [{card.get('msg_code', card.get('msg_name', 'N/A'))[:30]}] {card.get('msg_name', '')[:60]}")
    lines.append("━" * 67)
    lines.append("")
    lines.append(f"🎯 이 사례는 {card.get('team') or '담당팀 미정'}의 '{card.get('msg_name', '')}' 메시지를 검수 통과 형태로 개선한 정답지입니다.")
    lines.append("")

    lines.append("━━━ ASIS (직원 작성 원본) ━━━")
    lines.append(card.get("asis_marked", ""))
    lines.append("")

    lines.append("━━━ TOBE (검수자 개선안 — 인라인 마커) ━━━")
    lines.append(card.get("tobe_marked", ""))
    lines.append("")

    lines.append("━━━ 메타데이터 ━━━")
    lines.append(f"- 신뢰도 라벨: {card.get('label', 'LOW')}")
    lines.append(f"- 도메인: {card.get('domain', 'misc')}")
    if card.get("team"):
        lines.append(f"- 담당팀: {card['team']}")
    if card.get("msg_code"):
        lines.append(f"- 메시지 코드: {card['msg_code']}")
    if card.get("msg_name"):
        lines.append(f"- 메시지 이름: {card['msg_name']}")
    if card.get("agree"):
        lines.append(f"- 협의결과: {card['agree']}")
    if card.get("review"):
        lines.append(f"- 검토결과: {card['review']}")
    if card.get("feedback"):
        feedback_first = str(card["feedback"]).split("\n")[0][:50]
        lines.append(f"- 담당팀 피드백 첫 라인: {feedback_first}")
    # 시트 4 추가 메타
    if card.get("customer_confirm"):
        cc_short = str(card["customer_confirm"])[:80].replace("\n", " ")
        lines.append(f"- 고객사 컨펌의견: {cc_short}")
    if card.get("additional_suggest"):
        as_short = str(card["additional_suggest"])[:80].replace("\n", " ")
        lines.append(f"- 와이어링크 추가 제안(02/25): {as_short}")
    # 시트 2 추가 메타
    if card.get("wirelink_memo"):
        wm_short = str(card["wirelink_memo"])[:80].replace("\n", " ")
        lines.append(f"- 와이어링크 메모: {wm_short}")
    lines.append(f"- 원천: 정답데이터2 / {card.get('sheet')} R{card.get('row')}")
    lines.append("")

    return "\n".join(lines)


# ============== 표본 추출 ==============
def find_sample_blank_agree(wb, shared_strings):
    """1차 시트에서 협의결과=blank·TOBE 채워진 첫 행 찾기 (사용자 통찰 검증)"""
    sn = "1차_국내해외주식채권_피드백반영"
    cfg = SHEET_CFG[sn]
    ws = wb[sn]
    for r in range(3, ws.max_row + 1):
        agree = ws.cell(row=r, column=cfg["agree"]).value
        tobe = ws.cell(row=r, column=cfg["tobe"]).value
        if (not agree or not str(agree).strip()) and tobe and str(tobe).strip():
            return r
    return None


def extract_card_at(wb, shared_strings, sheet_name, row):
    """특정 행에서 카드 데이터 추출"""
    cfg = SHEET_CFG[sheet_name]
    ws = wb[sheet_name]

    def get(key):
        col = cfg.get(key)
        if not col:
            return None
        return ws.cell(row=row, column=col).value

    asis_segs = extract_cell_segments(XLSX_PATH, shared_strings, sheet_name, row, cfg["asis"])
    tobe_segs = extract_cell_segments(XLSX_PATH, shared_strings, sheet_name, row, cfg["tobe"])

    # 텍스트만 fallback (rich text 없을 때)
    if not asis_segs and get("asis"):
        asis_segs = [{"text": str(get("asis")), "color": "DEFAULT"}]
    if not tobe_segs and get("tobe"):
        tobe_segs = [{"text": str(get("tobe")), "color": "DEFAULT"}]

    label = classify_label(get("agree"), get("feedback"), get("review"))
    domain = classify_domain(get("team"), get("msg_name"), sheet_name)

    card = {
        "sheet": sheet_name, "row": row,
        "msg_code": str(get("msg_code") or "").strip(),
        "msg_name": str(get("msg_name") or "").strip(),
        "team": str(get("team") or "").strip(),
        "asis_marked": render_inline_marked(asis_segs),
        "tobe_marked": render_inline_marked(tobe_segs),
        "agree": str(get("agree") or "").strip(),
        "review": str(get("review") or "").strip(),
        "feedback": str(get("feedback") or "").strip(),
        "label": label,
        "domain": domain,
    }
    # 시트 4 추가
    if "4차" in sheet_name:
        card["customer_confirm"] = str(get("customer_confirm") or "").strip()
        card["additional_suggest"] = str(get("additional_suggest") or "").strip()
    # 시트 2 추가
    if "2차" in sheet_name:
        card["wirelink_memo"] = str(get("wirelink_memo") or "").strip()
    return card


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=== Phase A: 표본 5건 테스트 ===\n")
    print("[1/2] sharedStrings.xml 파싱 (rich text 색 정보)...")
    shared_strings = load_shared_strings_segments(XLSX_PATH)
    print(f"  shared strings: {len(shared_strings)}개\n")

    wb = load_workbook(XLSX_PATH, data_only=True)

    # 표본 1: 1차 LN102 (NEGATIVE 검증)
    samples = []

    # 1차 시트에서 LN102 찾기
    ws1 = wb["1차_국내해외주식채권_피드백반영"]
    for r in range(3, ws1.max_row + 1):
        if str(ws1.cell(row=r, column=2).value or "").strip() == "LN102":
            samples.append(("01_1차_LN102_NEGATIVE", "1차_국내해외주식채권_피드백반영", r))
            break

    # 표본 2: 1차 협의=blank·TOBE 채워진 행
    blank_row = find_sample_blank_agree(wb, shared_strings)
    if blank_row:
        samples.append(("02_1차_blank_TOBE채움_사용자통찰", "1차_국내해외주식채권_피드백반영", blank_row))

    # 표본 3: 2차 AC283
    ws2 = wb["2차_파생 등_피드백반영"]
    for r in range(3, ws2.max_row + 1):
        if str(ws2.cell(row=r, column=2).value or "").strip() == "AC283":
            samples.append(("03_2차_AC283_MEDIUM_와이어링크메모", "2차_파생 등_피드백반영", r))
            break

    # 표본 4: 3차 첫 행 (msg_code 없음)
    ws3 = wb["3차_프로세스 외_피드백 반영"]
    for r in range(3, ws3.max_row + 1):
        v = ws3.cell(row=r, column=4).value  # tobe
        if v and str(v).strip():
            samples.append(("04_3차_msg_code없음_process도메인", "3차_프로세스 외_피드백 반영", r))
            break

    # 표본 5: 4차 AP007
    ws4 = wb["4차_ 연금 전체"]
    for r in range(3, ws4.max_row + 1):
        if str(ws4.cell(row=r, column=2).value or "").strip() == "AP007":
            samples.append(("05_4차_AP007_pension_고객사컨펌", "4차_ 연금 전체", r))
            break

    print(f"[2/2] 표본 카드 생성 ({len(samples)}건)...\n")

    for i, (label, sheet, row) in enumerate(samples, start=1):
        try:
            card = extract_card_at(wb, shared_strings, sheet, row)
            text = make_card_text(card)
            output_file = OUTPUT_DIR / f"{label}.txt"
            output_file.write_text(text, encoding="utf-8")
            print(f"[{i}/{len(samples)}] {label}: 라벨={card['label']}, 도메인={card['domain']}, 글자수={len(text)}")
        except Exception as e:
            print(f"[{i}/{len(samples)}] {label}: 오류 - {e}")

    print(f"\n=== 표본 5건 생성 완료 → {OUTPUT_DIR} ===")
    print("사용자 검토 후 Phase B 진행 결정.")


if __name__ == "__main__":
    main()
