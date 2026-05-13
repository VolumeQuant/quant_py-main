"""
2단계: 5축 분류·라벨링 - 5건 샘플 검증

5축 신호:
  1. 협의결과 (반영/기존유지/추후반영/blank/자유텍스트/등)
  2. 담당팀 피드백 첫 라인 (수용/미수용/원안유지 등)
  3. 검토결과 (검토완료/검토필요/검수완료)
  4. rich text 색 분포 (DEFAULT/BLUE/RED 비율)
  5. 시트 (1차/2차/3차/4차)

5단계 라벨 자동 부여:
  - HIGH    : 반영 + 수용/이견없음 + 빨강 적음
  - MEDIUM  : 검토완료+blank or 단독 수용
  - LOW     : 검토필요+추후반영 or 자유텍스트 협의 (참조용)
  - NEGATIVE: 기존유지 or 미수용/원안유지
  - EVAL_ONLY: blank+blank (작업 미진행)

도메인 자동 분류 (팀명 → 도메인):
  - 연금업무개발팀 → pension (132건이라 IRP/DC/misc 3분할)
  - 증권운영팀 → credit_loan
  - 디지털영업팀·디지털마케팅팀·WM파생마케팅팀 → marketing
  - 상품서비스팀·상품개발팀·상품솔루션팀 → product
  - 파생Solution팀·신탁운용팀·Wrap솔루션팀·Equity → derivatives
  - 증권결제팀·예탁결제팀·글로벌주식솔루션팀 → settlement
  - 그 외 → misc

msg_code 중복 처리:
  - 같은 코드 여러 행 시 신뢰도 HIGH > MEDIUM > LOW 순
  - 동일 신뢰도면 더 최근 시트(시트 번호 높은 것) 우선
"""

import zipfile
import re
import xml.etree.ElementTree as ET
from collections import defaultdict, Counter
import openpyxl

XLSX_PATH = r"C:\dev\guide\정답데이터2\정답데이터2.xlsx"
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# ===== 1단계 함수 재사용 =====

def parse_shared_strings(xlsx_zip):
    with xlsx_zip.open("xl/sharedStrings.xml") as f:
        root = ET.parse(f).getroot()
    shared_list = []
    for si in root.findall("main:si", NS):
        segments = []
        runs = si.findall("main:r", NS)
        if runs:
            for r in runs:
                rpr = r.find("main:rPr", NS)
                color = "DEFAULT"
                if rpr is not None:
                    color_el = rpr.find("main:color", NS)
                    if color_el is not None and "rgb" in color_el.attrib:
                        color = color_el.attrib["rgb"]
                t_el = r.find("main:t", NS)
                text = t_el.text if t_el is not None and t_el.text else ""
                if text:
                    segments.append({"color": color, "text": text})
        else:
            t_el = si.find("main:t", NS)
            text = t_el.text if t_el is not None and t_el.text else ""
            if text:
                segments.append({"color": "DEFAULT", "text": text})
        shared_list.append(segments)
    return shared_list


def get_sheet_xml_paths(xlsx_zip):
    with xlsx_zip.open("xl/workbook.xml") as f:
        wb_root = ET.parse(f).getroot()
    with xlsx_zip.open("xl/_rels/workbook.xml.rels") as f:
        rels_root = ET.parse(f).getroot()
    rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    rid_to_target = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("r:Relationship", rels_ns)
    }
    sheet_map = {}
    sheets_el = wb_root.find("main:sheets", NS)
    rid_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    for sheet in sheets_el.findall("main:sheet", NS):
        name = sheet.attrib["name"]
        rid = sheet.attrib[rid_attr]
        target = rid_to_target[rid]
        full_path = "xl/" + target if not target.startswith("xl/") else target
        sheet_map[name] = full_path
    return sheet_map


def col_letter_to_idx(letter):
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_sheet_cells(xlsx_zip, sheet_path):
    with xlsx_zip.open(sheet_path) as f:
        root = ET.parse(f).getroot()
    cells = {}
    sheet_data = root.find("main:sheetData", NS)
    if sheet_data is None:
        return cells
    for row_el in sheet_data.findall("main:row", NS):
        row_num = int(row_el.attrib["r"])
        for cell_el in row_el.findall("main:c", NS):
            ref = cell_el.attrib["r"]
            m = re.match(r"^([A-Z]+)(\d+)$", ref)
            if not m:
                continue
            col_letter = m.group(1)
            col_idx = col_letter_to_idx(col_letter)
            t_attr = cell_el.attrib.get("t", "n")
            v_el = cell_el.find("main:v", NS)
            if v_el is None:
                continue
            if t_attr == "s":
                cells[(row_num, col_idx)] = ("shared", int(v_el.text))
            else:
                cells[(row_num, col_idx)] = ("inline", v_el.text or "")
    return cells


def get_rich_text(cells, shared_strings, row, col):
    cell_data = cells.get((row, col))
    if cell_data is None:
        return []
    kind, value = cell_data
    if kind == "shared":
        idx = value
        if 0 <= idx < len(shared_strings):
            return shared_strings[idx]
    elif kind == "inline":
        return [{"color": "DEFAULT", "text": str(value)}]
    return []


def normalize_color(rgb):
    if rgb in ("DEFAULT", "FF000000"):
        return "DEFAULT"
    red_set = {"FFFF0000", "FFCC0000", "FFFF00FF", "FF9900FF"}
    if rgb in red_set:
        return "RED"
    blue_set = {"FF0070C0", "FF4A86E8", "FF1155CC", "FF0000FF", "FF6AA84F"}
    if rgb in blue_set:
        return "BLUE"
    return "OTHER"


# ===== 2단계: 5축 분류 =====

# 시트별 컬럼 매핑 (영역 2에서 분석한 결과)
SHEET_CFG = {
    "1차_국내해외주식채권_피드백반영": {
        "asis": 8, "tobe": 9, "feedback": 10, "team": 14,
        "review": 18, "agree": 21, "code": 2, "name": 3, "remark": 17,
    },
    "2차_파생 등_피드백반영": {
        "asis": 8, "tobe": 9, "feedback": 19, "team": 12,
        "review": 16, "agree": 18, "code": 2, "name": 3, "remark": 15,
    },
    "3차_프로세스 외_피드백 반영": {
        "asis": 3, "tobe": 4, "feedback": 5, "team": 6,
        "review": None, "agree": 8, "code": None, "name": 2, "remark": 7,
    },
    "4차_ 연금 전체": {
        "asis": 8, "tobe": 9, "feedback": 17, "team": 12,
        "review": 14, "agree": 16, "code": 2, "name": 3, "remark": 13,
    },
}

# 팀 → 도메인 매핑
TEAM_DOMAIN_MAP = {
    "연금업무개발팀": "pension",
    "증권운영팀": "credit_loan",
    "디지털영업팀": "marketing",
    "디지털마케팅팀": "marketing",
    "WM파생마케팅팀": "marketing",
    "디지털상품마케팅팀": "marketing",
    "상품서비스팀": "product",
    "상품개발팀": "product",
    "상품솔루션팀": "product",
    "Sage컨설팅팀": "product",
    "파생Solution팀": "derivatives",
    "신탁운용팀": "derivatives",
    "Wrap솔루션팀": "derivatives",
    "Equity": "derivatives",
    "증권결제팀": "settlement",
    "예탁결제팀": "settlement",
    "글로벌주식솔루션팀": "settlement",
    "프로세스분석팀": "settlement",
    "채널서비스팀": "settlement",
    "AML솔루션팀": "settlement",
    "RP운용팀": "derivatives",
    "종금운용팀": "derivatives",
    "IPO2팀": "settlement",
    "해외채권상품운용팀": "misc",
    "개인투자용국채팀": "misc",
    "금융소비자보호팀": "misc",
    # 그 외는 misc
}

# 연금 하위 분류 (메시지명에서 추출)
PENSION_SUB_KEYWORDS = {
    "irp": ["IRP", "개인형"],
    "dc": ["DC", "확정기여"],
    # 그 외 연금은 misc
}

POSITIVE_FEEDBACK = {"수용", "이견없음", "이의없음", "이상없음", "개선안 TOBE 동일", "시용", "진행요청"}
NEGATIVE_FEEDBACK = {"미수용", "원안유지"}


def classify_team_to_domain(team_str):
    if not team_str:
        return "misc"
    team_clean = re.sub(r"\d+", "", str(team_str)).strip().split()[0][:30]
    # 우선 정확 매칭
    for key, domain in TEAM_DOMAIN_MAP.items():
        if key in team_clean:
            return domain
    return "misc"


def classify_pension_sub(msg_name):
    """연금 하위 분류 (IRP/DC/misc)."""
    if not msg_name:
        return "misc"
    name_str = str(msg_name)
    for sub, keywords in PENSION_SUB_KEYWORDS.items():
        for kw in keywords:
            if kw in name_str:
                return sub
    return "misc"


def color_distribution(segments):
    """ASIS 또는 TOBE의 색별 글자 수 비율 반환."""
    counter = defaultdict(int)
    for s in segments:
        counter[normalize_color(s["color"])] += len(s["text"])
    total = sum(counter.values())
    if total == 0:
        return {}
    return {k: v / total for k, v in counter.items()}


def first_line(text):
    """텍스트 첫 라인 추출 (담당팀 피드백 분류용)."""
    if not text:
        return ""
    return str(text).strip().split("\n")[0].strip()[:30]


def assign_5axis_label(sheet_name, agree, feedback_first, review, tobe_color_dist):
    """5축 신호로 5단계 라벨 부여."""
    blue_pct = tobe_color_dist.get("BLUE", 0)
    red_pct = tobe_color_dist.get("RED", 0)

    # 시트별 미진행 케이스
    if sheet_name == "4차_ 연금 전체":
        # 4차는 검수완료/검토완료가 의미 있는 신호
        if review in ("검수완료", "검토완료"):
            return "MEDIUM"
        return "EVAL_ONLY"

    # NEGATIVE 우선 검사
    if agree == "기존유지" or feedback_first in NEGATIVE_FEEDBACK:
        return "NEGATIVE"

    if agree == "삭제":
        return "NEGATIVE"

    # HIGH: 반영 + (수용/이견없음 etc)
    if agree == "반영" and feedback_first in POSITIVE_FEEDBACK:
        return "HIGH"
    if agree == "반영":  # 반영만 있어도 HIGH (피드백 없을 수도)
        return "HIGH"

    # MEDIUM: 검토완료 OR 단독 수용 OR 3차 검토완료
    if review in ("검토완료", "검수완료"):
        return "MEDIUM"
    if feedback_first in POSITIVE_FEEDBACK:
        return "MEDIUM"

    # LOW: 검토필요·추후반영 or 자유텍스트 협의 (참조용)
    if agree == "추후반영":
        return "LOW"
    if agree and len(agree) > 30:  # 자유 텍스트
        return "LOW"
    if review == "검토필요":
        return "LOW"

    # 그 외 미진행
    return "EVAL_ONLY"


def select_5_samples(wb):
    """1단계와 동일한 5건 샘플."""
    samples = []
    sn = "1차_국내해외주식채권_피드백반영"
    cfg = SHEET_CFG[sn]
    ws = wb[sn]

    # 1) 1차 기존유지 첫 건
    for r in range(3, ws.max_row + 1):
        agree = str(ws.cell(r, cfg["agree"]).value or "").strip()
        asis = ws.cell(r, cfg["asis"]).value
        tobe = ws.cell(r, cfg["tobe"]).value
        if agree == "기존유지" and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn, r, "기존유지(NEGATIVE)"))
            break

    # 2) 1차 검토완료+blank 첫 건
    for r in range(3, ws.max_row + 1):
        review = str(ws.cell(r, cfg["review"]).value or "").strip()
        agree = str(ws.cell(r, cfg["agree"]).value or "").strip()
        asis = ws.cell(r, cfg["asis"]).value
        tobe = ws.cell(r, cfg["tobe"]).value
        if review == "검토완료" and not agree and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn, r, "검토완료+blank(MEDIUM)"))
            break

    # 3) 2차 자유 텍스트
    sn2 = "2차_파생 등_피드백반영"
    cfg2 = SHEET_CFG[sn2]
    ws2 = wb[sn2]
    for r in range(3, ws2.max_row + 1):
        agree = str(ws2.cell(r, cfg2["agree"]).value or "").strip()
        asis = ws2.cell(r, cfg2["asis"]).value
        tobe = ws2.cell(r, cfg2["tobe"]).value
        if agree and agree not in ("반영", "추후반영", "기존유지", "삭제") and len(agree) > 30:
            if asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
                samples.append((sn2, r, "자유 텍스트 협의(LOW)"))
                break

    # 4) 4차 검수완료
    sn4 = "4차_ 연금 전체"
    cfg4 = SHEET_CFG[sn4]
    ws4 = wb[sn4]
    for r in range(3, ws4.max_row + 1):
        review = str(ws4.cell(r, cfg4["review"]).value or "").strip()
        asis = ws4.cell(r, cfg4["asis"]).value
        tobe = ws4.cell(r, cfg4["tobe"]).value
        if review in ("검수완료", "검토완료") and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn4, r, "4차 연금 검수완료(MEDIUM)"))
            break

    # 5) 1차 R6 (rainbow — 빨강+파랑)
    samples.append(("1차_국내해외주식채권_피드백반영", 6, "빨강+파랑(MEDIUM 또는 LOW)"))

    return samples


def main():
    print(f"=== 2단계: 5축 분류·라벨링 5건 샘플 검증 ===\n")

    with zipfile.ZipFile(XLSX_PATH) as z:
        shared_strings = parse_shared_strings(z)
        sheet_map = get_sheet_xml_paths(z)
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        samples = select_5_samples(wb)

        for sn, row, sample_desc in samples:
            cells = parse_sheet_cells(z, sheet_map[sn])
            cfg = SHEET_CFG[sn]
            ws = wb[sn]

            # 5축 신호 추출
            asis_segs = get_rich_text(cells, shared_strings, row, cfg["asis"])
            tobe_segs = get_rich_text(cells, shared_strings, row, cfg["tobe"])

            agree_raw = ws.cell(row, cfg["agree"]).value if cfg["agree"] else None
            agree = str(agree_raw or "").strip()
            agree_short = agree[:40] if agree else "(blank)"

            review_raw = ws.cell(row, cfg["review"]).value if cfg["review"] else None
            review = str(review_raw or "").strip()

            feedback_raw = ws.cell(row, cfg["feedback"]).value if cfg["feedback"] else None
            feedback_first_line = first_line(feedback_raw)

            tobe_color_dist = color_distribution(tobe_segs)

            team_raw = ws.cell(row, cfg["team"]).value if cfg["team"] else None
            team = str(team_raw or "").strip().split("\n")[0]

            msg_code = ws.cell(row, cfg["code"]).value if cfg["code"] else None
            msg_name = ws.cell(row, cfg["name"]).value if cfg["name"] else None

            # 분류
            label = assign_5axis_label(sn, agree, feedback_first_line, review, tobe_color_dist)
            domain = classify_team_to_domain(team)
            if domain == "pension":
                pension_sub = classify_pension_sub(msg_name)
                domain_full = f"pension_{pension_sub}"
            else:
                domain_full = domain

            # 출력
            print(f"{'='*80}")
            print(f"[{sample_desc}] {sn} R{row}")
            print(f"  msg_code: {msg_code} | msg_name: {msg_name}")
            print(f"  team: {team}")
            print(f"  ──── 5축 신호 ────")
            print(f"  ① 협의결과:    {agree_short}")
            print(f"  ② 담당팀 피드백 첫 라인: {feedback_first_line if feedback_first_line else '(blank)'}")
            print(f"  ③ 검토결과:    {review or '(blank)'}")
            print(f"  ④ TOBE 색 분포: {' '.join([f'{k}={v*100:.0f}%' for k,v in tobe_color_dist.items()])}")
            print(f"  ⑤ 시트:       {sn}")
            print(f"  ──── 자동 라벨 ────")
            print(f"  🏷️  신뢰도 라벨: {label}")
            print(f"  📂 도메인:     {domain_full}")
            print()


if __name__ == "__main__":
    main()
