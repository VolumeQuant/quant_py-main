"""
2단계 보강: 전체 409건 5축 분류 분포 측정

목적: 5건 샘플이 잘 작동하는 것 확인. 이제 전체에 적용했을 때
어떤 라벨·도메인 분포가 나오는지 보고 카드 docx 분할이 합리적인지 검증.
"""

import zipfile
import re
import xml.etree.ElementTree as ET
from collections import defaultdict, Counter
import openpyxl

# 이전 스크립트 import 대신 같이 정의 (단일 파일 실행 편의)
import sys
sys.path.insert(0, r"C:\dev\guide\tools")
from classify_5axis_sample import (
    parse_shared_strings, get_sheet_xml_paths, parse_sheet_cells,
    get_rich_text, normalize_color, color_distribution, first_line,
    classify_team_to_domain, classify_pension_sub, assign_5axis_label,
    SHEET_CFG, XLSX_PATH,
)


def main():
    print("=== 전체 409건 5축 분류 분포 측정 ===\n")

    label_counter = Counter()
    domain_counter = Counter()
    domain_label_cross = defaultdict(Counter)
    msg_code_counter = Counter()
    pension_sub_counter = Counter()

    duplicate_codes = []  # (sheet, row, code) 중복 추적
    code_to_rows = defaultdict(list)

    with zipfile.ZipFile(XLSX_PATH) as z:
        shared_strings = parse_shared_strings(z)
        sheet_map = get_sheet_xml_paths(z)
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

        target_sheets = [
            "1차_국내해외주식채권_피드백반영",
            "2차_파생 등_피드백반영",
            "4차_ 연금 전체",
            # 3차는 컬럼 시맨틱 다름. 일단 제외하고 나중에 수동 처리.
        ]

        total = 0
        for sn in target_sheets:
            cells = parse_sheet_cells(z, sheet_map[sn])
            cfg = SHEET_CFG[sn]
            ws = wb[sn]
            for r in range(3, ws.max_row + 1):
                asis = ws.cell(r, cfg["asis"]).value
                tobe = ws.cell(r, cfg["tobe"]).value
                if not (asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30):
                    continue
                total += 1

                # 5축 추출
                tobe_segs = get_rich_text(cells, shared_strings, r, cfg["tobe"])
                agree = str(ws.cell(r, cfg["agree"]).value or "").strip() if cfg["agree"] else ""
                review = str(ws.cell(r, cfg["review"]).value or "").strip() if cfg["review"] else ""
                feedback = first_line(ws.cell(r, cfg["feedback"]).value if cfg["feedback"] else None)
                team = str(ws.cell(r, cfg["team"]).value or "").strip().split("\n")[0] if cfg["team"] else ""
                msg_code = str(ws.cell(r, cfg["code"]).value or "").strip() if cfg["code"] else ""
                msg_name = ws.cell(r, cfg["name"]).value if cfg["name"] else ""

                # 라벨링
                tobe_color_dist = color_distribution(tobe_segs)
                label = assign_5axis_label(sn, agree, feedback, review, tobe_color_dist)
                domain = classify_team_to_domain(team)
                if domain == "pension":
                    sub = classify_pension_sub(msg_name)
                    pension_sub_counter[sub] += 1
                    domain_full = f"pension_{sub}"
                else:
                    domain_full = domain

                label_counter[label] += 1
                domain_counter[domain_full] += 1
                domain_label_cross[domain_full][label] += 1
                if msg_code:
                    msg_code_counter[msg_code] += 1
                    code_to_rows[msg_code].append((sn, r, label))

        print(f"전체 유효 페어: {total}건\n")

        print("【5단계 라벨 분포】")
        for label in ["HIGH", "MEDIUM", "LOW", "NEGATIVE", "EVAL_ONLY"]:
            cnt = label_counter[label]
            pct = cnt / total * 100 if total else 0
            print(f"  {label:12s}: {cnt:4d}건 ({pct:.1f}%)")
        print()

        print("【도메인 분포】")
        for domain, cnt in domain_counter.most_common():
            pct = cnt / total * 100
            print(f"  {domain:20s}: {cnt:4d}건 ({pct:.1f}%)")
        print()

        print("【도메인 × 라벨 교차】")
        print(f"  {'domain':<20s} {'HIGH':>5s} {'MED':>5s} {'LOW':>5s} {'NEG':>5s} {'EVAL':>5s}")
        for domain, _ in domain_counter.most_common():
            row = domain_label_cross[domain]
            print(f"  {domain:<20s} "
                  f"{row['HIGH']:>5d} {row['MEDIUM']:>5d} {row['LOW']:>5d} "
                  f"{row['NEGATIVE']:>5d} {row['EVAL_ONLY']:>5d}")
        print()

        # 카드용 (HIGH+MEDIUM+LOW)와 negative 별도 집계
        cards_per_domain = {}
        negatives_per_domain = {}
        for domain, sub in domain_label_cross.items():
            cards_per_domain[domain] = sub["HIGH"] + sub["MEDIUM"] + sub["LOW"]
            negatives_per_domain[domain] = sub["NEGATIVE"]

        print("【docx 분할 시뮬레이션】 (HIGH+MEDIUM+LOW = 카드)")
        total_cards = sum(cards_per_domain.values())
        total_negatives = sum(negatives_per_domain.values())
        print(f"  총 카드 수: {total_cards}건")
        print(f"  총 negative: {total_negatives}건 (cx_exceptions_misuyong.docx)")
        print(f"  EVAL_ONLY (vector DB 미반영): {label_counter['EVAL_ONLY']}건\n")

        for domain, n_cards in sorted(cards_per_domain.items(), key=lambda x: -x[1]):
            n_neg = negatives_per_domain.get(domain, 0)
            print(f"  cx_goldens2_{domain:<18s}: {n_cards:3d} 카드  (negative {n_neg}건은 별도)")
        print()

        # msg_code 중복
        dup_codes = [(c, n) for c, n in msg_code_counter.items() if n > 1]
        print(f"【msg_code 중복】 {len(dup_codes)}개 코드 중복 (총 {sum(n-1 for c,n in dup_codes)}건 추가 발생)")
        for code, n in sorted(dup_codes, key=lambda x: -x[1])[:10]:
            rows = code_to_rows[code]
            labels = [r[2] for r in rows]
            print(f"  {code}: {n}회 — labels={labels}")


if __name__ == "__main__":
    main()
