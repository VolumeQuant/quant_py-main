"""
1단계: 정답데이터2.xlsx에서 글자 색까지 정확히 추출 — 5건 샘플 검증

목적:
  사용자가 결정한 카드 형식(검정/파랑/빨강 3계층 보존)을 위해
  xlsx의 rich text 색 정보를 정확히 segment로 분리해서 뽑아내는 것이
  기술적으로 가능한지 5건으로 먼저 검증한다.

방식:
  xlsx는 ZIP. 안의 xl/sharedStrings.xml과 xl/worksheets/sheetN.xml을
  직접 파싱해서 셀별 rich text segment를 [{color, text}] 리스트로 추출.

샘플 5건 선정 (다양성):
  1. 1차 시트 협의결과='기존유지' 첫 건 (negative 라벨링 검증)
  2. 1차 시트 검토결과='검토완료' AND 협의결과=blank 첫 건 (가장 큰 덩어리)
  3. 2차 시트 자유 텍스트 협의결과 첫 건 (특이 케이스)
  4. 4차(연금) 검수완료 첫 건 (다른 작업 단계)
  5. 빨강+파랑 둘 다 있는 사례 첫 건 (3계층 정보 풍부)
"""

import zipfile
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
import openpyxl

XLSX_PATH = r"C:\dev\guide\정답데이터2\정답데이터2.xlsx"

# Excel 네임스페이스
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def parse_shared_strings(xlsx_zip):
    """sharedStrings.xml을 파싱해서 인덱스별 rich text segment 리스트 반환.

    각 항목 = [{"color": "FF000000", "text": "..."}, ...]
    색이 명시 안 된 segment는 color="DEFAULT".
    """
    with xlsx_zip.open("xl/sharedStrings.xml") as f:
        tree = ET.parse(f)
    root = tree.getroot()

    shared_list = []
    for si in root.findall("main:si", NS):
        segments = []

        # <r> 블록이 있으면 rich text. 없으면 단순 <t>
        runs = si.findall("main:r", NS)
        if runs:
            for r in runs:
                rpr = r.find("main:rPr", NS)
                color = "DEFAULT"
                if rpr is not None:
                    color_el = rpr.find("main:color", NS)
                    if color_el is not None and "rgb" in color_el.attrib:
                        color = color_el.attrib["rgb"]
                t_el = r.find("main:t", NS)
                text = t_el.text if t_el is not None and t_el.text else ""
                if text:
                    segments.append({"color": color, "text": text})
        else:
            t_el = si.find("main:t", NS)
            text = t_el.text if t_el is not None and t_el.text else ""
            if text:
                segments.append({"color": "DEFAULT", "text": text})

        shared_list.append(segments)
    return shared_list


def get_sheet_xml_paths(xlsx_zip):
    """workbook.xml에서 시트 이름 → sheet xml 경로 매핑 반환."""
    # workbook.xml에서 시트 ID
    with xlsx_zip.open("xl/workbook.xml") as f:
        wb_root = ET.parse(f).getroot()

    # workbook.xml.rels로 ID → 파일 경로 매핑
    with xlsx_zip.open("xl/_rels/workbook.xml.rels") as f:
        rels_root = ET.parse(f).getroot()
    rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    rid_to_target = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("r:Relationship", rels_ns)
    }

    sheet_map = {}
    sheets_el = wb_root.find("main:sheets", NS)
    rid_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    for sheet in sheets_el.findall("main:sheet", NS):
        name = sheet.attrib["name"]
        rid = sheet.attrib[rid_attr]
        target = rid_to_target[rid]
        # target은 worksheets/sheet1.xml 형태. xl/ 경로 추가 필요
        full_path = "xl/" + target if not target.startswith("xl/") else target
        sheet_map[name] = full_path
    return sheet_map


def col_letter_to_idx(letter):
    """A=1, B=2, ..., Z=26, AA=27 변환."""
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_sheet_cells(xlsx_zip, sheet_path):
    """sheet xml을 파싱해서 (row, col) → sharedString 인덱스 또는 inline 값 매핑."""
    with xlsx_zip.open(sheet_path) as f:
        root = ET.parse(f).getroot()

    cells = {}
    sheet_data = root.find("main:sheetData", NS)
    if sheet_data is None:
        return cells

    for row_el in sheet_data.findall("main:row", NS):
        row_num = int(row_el.attrib["r"])
        for cell_el in row_el.findall("main:c", NS):
            ref = cell_el.attrib["r"]  # 예: "H3"
            m = re.match(r"^([A-Z]+)(\d+)$", ref)
            if not m:
                continue
            col_letter, _ = m.group(1), m.group(2)
            col_idx = col_letter_to_idx(col_letter)

            t_attr = cell_el.attrib.get("t", "n")  # 기본은 number
            v_el = cell_el.find("main:v", NS)
            if v_el is None:
                continue

            if t_attr == "s":
                # sharedString 인덱스
                cells[(row_num, col_idx)] = ("shared", int(v_el.text))
            else:
                # inline 또는 number/string
                cells[(row_num, col_idx)] = ("inline", v_el.text or "")

    return cells


def get_rich_text(cells, shared_strings, row, col):
    """셀의 rich text segment 리스트 반환. 없으면 빈 리스트."""
    cell_data = cells.get((row, col))
    if cell_data is None:
        return []

    kind, value = cell_data
    if kind == "shared":
        idx = value
        if 0 <= idx < len(shared_strings):
            return shared_strings[idx]
    elif kind == "inline":
        return [{"color": "DEFAULT", "text": str(value)}]
    return []


def normalize_color(rgb):
    """rgb 색을 분류 라벨로 정규화.
    빨강 계열 → RED, 파랑 계열 → BLUE, 검정/기본 → DEFAULT
    """
    if rgb in ("DEFAULT", "FF000000"):
        return "DEFAULT"

    # 빨강 계열: FF로 시작 (높은 R), 낮은 G·B
    # FFFF0000(빨강), FFCC0000(어두운 빨강), FFFF00FF(자홍?)
    red_set = {"FFFF0000", "FFCC0000", "FFFF00FF", "FF9900FF"}
    if rgb in red_set:
        return "RED"

    # 파랑 계열: 다양한 파랑·청록
    blue_set = {"FF0070C0", "FF4A86E8", "FF1155CC", "FF0000FF", "FF6AA84F"}
    if rgb in blue_set:
        return "BLUE"

    # 그 외는 OTHER 라벨로 표시
    return f"OTHER({rgb})"


def select_5_samples(wb):
    """다양성 있는 5건 샘플 선정. 행 좌표 (sheet_name, row, asis_col, tobe_col, label) 반환."""
    samples = []

    # 시트 컬럼 매핑
    sheet_cfg = {
        "1차_국내해외주식채권_피드백반영": {"asis": 8, "tobe": 9, "review": 18, "agree": 21},
        "2차_파생 등_피드백반영": {"asis": 8, "tobe": 9, "review": 16, "agree": 18},
        "4차_ 연금 전체": {"asis": 8, "tobe": 9, "review": 14, "agree": 16},
    }

    # 1) 1차 협의결과='기존유지' 첫 건
    sn = "1차_국내해외주식채권_피드백반영"
    cfg = sheet_cfg[sn]
    ws = wb[sn]
    for r in range(3, ws.max_row + 1):
        agree = str(ws.cell(r, cfg["agree"]).value or "").strip()
        asis = ws.cell(r, cfg["asis"]).value
        tobe = ws.cell(r, cfg["tobe"]).value
        if agree == "기존유지" and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn, r, cfg["asis"], cfg["tobe"], "★ 1차 기존유지(NEGATIVE)"))
            break

    # 2) 1차 검토완료+blank 첫 건
    for r in range(3, ws.max_row + 1):
        review = str(ws.cell(r, cfg["review"]).value or "").strip()
        agree = str(ws.cell(r, cfg["agree"]).value or "").strip()
        asis = ws.cell(r, cfg["asis"]).value
        tobe = ws.cell(r, cfg["tobe"]).value
        if review == "검토완료" and not agree and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn, r, cfg["asis"], cfg["tobe"], "★ 1차 검토완료+blank(MEDIUM)"))
            break

    # 3) 2차 자유 텍스트 협의결과 첫 건
    sn2 = "2차_파생 등_피드백반영"
    cfg2 = sheet_cfg[sn2]
    ws2 = wb[sn2]
    for r in range(3, ws2.max_row + 1):
        agree = str(ws2.cell(r, cfg2["agree"]).value or "").strip()
        asis = ws2.cell(r, cfg2["asis"]).value
        tobe = ws2.cell(r, cfg2["tobe"]).value
        # 자유 텍스트 = 라벨 4종(반영/추후반영/기존유지/삭제) 외이고 길이 30자 이상
        if agree and agree not in ("반영", "추후반영", "기존유지", "삭제") and len(agree) > 30:
            if asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
                samples.append((sn2, r, cfg2["asis"], cfg2["tobe"], "★ 2차 자유 텍스트 협의(특이 케이스)"))
                break

    # 4) 4차(연금) 검수완료 첫 건
    sn4 = "4차_ 연금 전체"
    cfg4 = sheet_cfg[sn4]
    ws4 = wb[sn4]
    for r in range(3, ws4.max_row + 1):
        review = str(ws4.cell(r, cfg4["review"]).value or "").strip()
        asis = ws4.cell(r, cfg4["asis"]).value
        tobe = ws4.cell(r, cfg4["tobe"]).value
        if review in ("검수완료", "검토완료") and asis and tobe and len(str(asis)) > 30 and len(str(tobe)) > 30:
            samples.append((sn4, r, cfg4["asis"], cfg4["tobe"], "★ 4차 연금 검수완료(다른 작업 단계)"))
            break

    # 5) 빨강+파랑 둘 다 있는 사례 (1차/2차에서 ASIS·TOBE에 색 다양성)
    # 이건 segment 추출 후에야 알 수 있어서 별도 함수로 처리. 일단 자리만 잡아둠.
    samples.append(("__FIND_RAINBOW__", None, None, None, "★ 빨강+파랑 둘 다 있는 사례"))

    return samples


def find_rainbow_sample(xlsx_zip, shared_strings, sheet_map, sheet_cfg):
    """빨강+파랑 둘 다 ASIS 또는 TOBE에 있는 첫 사례 찾기."""
    for sn in ["1차_국내해외주식채권_피드백반영", "2차_파생 등_피드백반영"]:
        if sn not in sheet_map:
            continue
        cells = parse_sheet_cells(xlsx_zip, sheet_map[sn])
        cfg = sheet_cfg[sn]
        # row 범위 추정 (1차=145, 2차=132 정도)
        for r in range(3, 200):
            tobe_segs = get_rich_text(cells, shared_strings, r, cfg["tobe"])
            if not tobe_segs:
                continue
            colors = set(normalize_color(s["color"]) for s in tobe_segs)
            if "RED" in colors and "BLUE" in colors:
                return (sn, r, cfg["asis"], cfg["tobe"])
    return None


def render_sample(sample_label, sheet_name, row, asis_col, tobe_col, cells, shared_strings, ws):
    """한 샘플의 ASIS·TOBE 색 segment 출력."""
    print(f"\n{'='*80}")
    print(f"{sample_label} — 시트: {sheet_name}, Row: {row}")
    print(f"{'='*80}")

    # 추가 메타
    msg_code = ws.cell(row, 2).value if sheet_name != "3차_프로세스 외_피드백 반영" else None
    msg_name = ws.cell(row, 3).value if sheet_name != "3차_프로세스 외_피드백 반영" else None
    print(f"메시지 코드: {msg_code}")
    print(f"메시지 이름: {msg_name}")

    for col_label, col_idx in [("ASIS", asis_col), ("TOBE", tobe_col)]:
        segs = get_rich_text(cells, shared_strings, row, col_idx)
        print(f"\n--- {col_label} 셀 ({len(segs)} segment) ---")
        if not segs:
            print("(빈 셀)")
            continue
        # 색별 통계
        color_counter = defaultdict(int)
        for s in segs:
            color_counter[normalize_color(s["color"])] += len(s["text"])
        total = sum(color_counter.values())
        print(f"색 분포: ", end="")
        for c, cnt in sorted(color_counter.items(), key=lambda x: -x[1]):
            pct = cnt / total * 100 if total else 0
            print(f"{c}={cnt}자({pct:.0f}%) ", end="")
        print()
        # segment별 출력 (앞 1500자만)
        cumulative = 0
        for i, s in enumerate(segs):
            if cumulative > 1500:
                print(f"  ... (이하 생략)")
                break
            color_label = normalize_color(s["color"])
            text_preview = s["text"][:200].replace("\n", " ⏎ ")
            print(f"  [{color_label}] {text_preview}")
            cumulative += len(s["text"])


def main():
    print(f"=== 1단계: 정답데이터2 rich text 색 추출 5건 샘플 검증 ===")
    print(f"대상 파일: {XLSX_PATH}\n")

    # ZIP 파싱
    with zipfile.ZipFile(XLSX_PATH) as z:
        shared_strings = parse_shared_strings(z)
        print(f"[OK] sharedStrings.xml: {len(shared_strings)} 항목 로드")

        sheet_map = get_sheet_xml_paths(z)
        print(f"[OK] 시트 매핑: {len(sheet_map)} 시트")

        sheet_cfg = {
            "1차_국내해외주식채권_피드백반영": {"asis": 8, "tobe": 9, "review": 18, "agree": 21},
            "2차_파생 등_피드백반영": {"asis": 8, "tobe": 9, "review": 16, "agree": 18},
            "4차_ 연금 전체": {"asis": 8, "tobe": 9, "review": 14, "agree": 16},
        }

        # openpyxl로 메타 컬럼 읽기 (msg_code, msg_name 등)
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

        # 샘플 5건 선정
        samples = select_5_samples(wb)

        # 5번째 (rainbow) 사례는 별도 검색
        rainbow = find_rainbow_sample(z, shared_strings, sheet_map, sheet_cfg)
        if rainbow:
            samples[-1] = (rainbow[0], rainbow[1], rainbow[2], rainbow[3], "★ 빨강+파랑 둘 다 있는 사례")
        else:
            samples = samples[:-1]  # 못 찾으면 빼기
            print("⚠️  빨강+파랑 둘 다 있는 사례 못 찾음")

        # 각 샘플 셀 데이터 캐싱
        sheet_cells_cache = {}
        for s in samples:
            sn = s[0]
            if sn not in sheet_cells_cache:
                sheet_cells_cache[sn] = parse_sheet_cells(z, sheet_map[sn])
            print(f"  - {s[4]}: {sn} R{s[1]}")

        # 5건 렌더링
        for sample in samples:
            sn, row, asis_col, tobe_col, label = sample
            cells = sheet_cells_cache[sn]
            ws = wb[sn]
            render_sample(label, sn, row, asis_col, tobe_col, cells, shared_strings, ws)

    print(f"\n{'='*80}")
    print(f"검증 완료. 색 추출이 정상 작동하는지 위 결과를 확인해 주세요.")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
