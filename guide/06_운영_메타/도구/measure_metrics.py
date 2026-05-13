"""
v5.21 효과 측정 메트릭 5종 (메모리 Phase 7·8 5주차 결정)

메트릭:
1. TOBE BLEU: 0.4 → 0.65+
2. 해요체 비율 (행동 요청): <10% → >50%
3. ☎ 제거율: 95% → 100%
4. 형식적 인사 출현: 30% → <5%
5. VOC 4분기 정확도: 25% → 75%+

입력: baseline_v5.20_outputs_template.xlsx (사용자가 v5.20·v5.21 출력 채워 넣은 파일)
출력: output/평가셋/metrics_report.xlsx (v5.20·v5.21 메트릭 비교)
"""

import re
from pathlib import Path
from openpyxl import load_workbook, Workbook

EVAL_DIR = Path(r"C:\dev\guide\output\평가셋")

# 형식적 인사 키워드 (가이드 위반 표현)
FORMAL_GREETING_KEYWORDS = [
    "편안한 하루",
    "환절기 건강",
    "항상 감사",
    "늘 감사",
    "건강 유의",
    "건강 조심",
    "행복한 하루",
]


def measure_haeyo_ratio(text):
    """해요체 비율: 행동 요청 문장 중 '~해 주세요' 비율"""
    if not text:
        return 0.0
    sentences = re.split(r"[.!?\n]+", text)
    action = [
        s for s in sentences
        if any(kw in s for kw in ["주세요", "주십시오", "바랍니다", "하십시오"])
    ]
    if not action:
        return 0.0
    haeyo = [s for s in action if "주세요" in s and "주십시오" not in s]
    return len(haeyo) / len(action)


def measure_phone_removal(text):
    """☎ 제거: 0이면 제거됨 (1), 있으면 (0)"""
    return 1 if text and "☎" not in text else 0


def measure_formal_greeting(text):
    """형식적 인사 출현 (1=출현, 0=미출현)"""
    if not text:
        return 0
    return 1 if any(kw in text for kw in FORMAL_GREETING_KEYWORDS) else 0


def measure_voc_classify(asis, hypothesis_type):
    """VOC 4분기 정확도 — ASIS 신호로 시나리오 추론, hypothesis와 일치 여부"""
    if not asis:
        return None
    expected = "VOC"  # 기본
    if any(kw in asis for kw in ["라고 알고", "인 줄 알", "라고 해서"]):
        expected = "VOC_misclaim"
    elif any(kw in asis for kw in ["거듭", "대단히", "진심으로"]) and any(
        kw in asis for kw in ["금감원", "언론", "법적"]
    ):
        expected = "VOC_edit"
    elif any(kw in asis for kw in ["처리 어렵", "지원 안", "처리 불가"]):
        expected = "VOC_reject"
    return 1 if expected == hypothesis_type else 0


def simple_bleu(reference, hypothesis):
    """sacrebleu 없을 때 단순 unigram overlap (대체 메트릭)"""
    if not reference or not hypothesis:
        return 0.0
    ref_tokens = set(reference.split())
    hyp_tokens = set(hypothesis.split())
    if not ref_tokens:
        return 0.0
    overlap = ref_tokens & hyp_tokens
    return len(overlap) / len(ref_tokens)


def measure_bleu(reference, hypothesis):
    """TOBE BLEU 측정 — sacrebleu 있으면 사용, 없으면 unigram overlap"""
    try:
        import sacrebleu

        return sacrebleu.sentence_bleu(hypothesis, [reference]).score / 100
    except ImportError:
        return simple_bleu(reference, hypothesis)


def main():
    template_path = EVAL_DIR / "baseline_v5.20_outputs_template.xlsx"
    if not template_path.exists():
        print(f"ERR: {template_path} 없음. make_eval_set.py 먼저 실행")
        return

    wb = load_workbook(str(template_path), data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers, start=1) if h}

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i).value for h, i in col.items()}
        rows.append(row)

    metrics_v520 = {
        "haeyo_ratio": [],
        "phone_removal": [],
        "formal_greeting": [],
        "bleu": [],
    }
    metrics_v521 = {
        "haeyo_ratio": [],
        "phone_removal": [],
        "formal_greeting": [],
        "bleu": [],
    }

    for r in rows:
        asis = r.get("asis", "") or ""
        tobe = r.get("tobe_reference", "") or ""
        v520 = r.get("v5.20_output", "") or ""
        v521 = r.get("v5.21_output", "") or ""

        if v520 and v520 != "[채워 넣기]":
            metrics_v520["haeyo_ratio"].append(measure_haeyo_ratio(v520))
            metrics_v520["phone_removal"].append(measure_phone_removal(v520))
            metrics_v520["formal_greeting"].append(measure_formal_greeting(v520))
            metrics_v520["bleu"].append(measure_bleu(tobe, v520))

        if v521 and v521 != "[채워 넣기]":
            metrics_v521["haeyo_ratio"].append(measure_haeyo_ratio(v521))
            metrics_v521["phone_removal"].append(measure_phone_removal(v521))
            metrics_v521["formal_greeting"].append(measure_formal_greeting(v521))
            metrics_v521["bleu"].append(measure_bleu(tobe, v521))

    def avg(arr):
        return sum(arr) / len(arr) if arr else 0.0

    print("\n=== 메트릭 비교 ===\n")
    print(f"  평가 사례 수: {len(rows)}")
    print(f"  v5.20 채워진: {len(metrics_v520['bleu'])} / v5.21 채워진: {len(metrics_v521['bleu'])}")
    print()
    print(f"  TOBE BLEU            v5.20: {avg(metrics_v520['bleu']):.3f}    v5.21: {avg(metrics_v521['bleu']):.3f}    (목표: 0.4 → 0.65+)")
    print(f"  해요체 비율 (행동 요청) v5.20: {avg(metrics_v520['haeyo_ratio']):.1%}    v5.21: {avg(metrics_v521['haeyo_ratio']):.1%}    (목표: <10% → >50%)")
    print(f"  ☎ 제거율             v5.20: {avg(metrics_v520['phone_removal']):.1%}    v5.21: {avg(metrics_v521['phone_removal']):.1%}    (목표: 95% → 100%)")
    print(f"  형식적 인사 출현률   v5.20: {avg(metrics_v520['formal_greeting']):.1%}    v5.21: {avg(metrics_v521['formal_greeting']):.1%}    (목표: 30% → <5%)")
    print()

    # 보고서 저장
    report_wb = Workbook()
    rs = report_wb.active
    rs.title = "metrics"
    rs.append(["metric", "v5.20", "v5.21", "target_v5.21", "delta"])
    rs.append(["TOBE BLEU", avg(metrics_v520["bleu"]), avg(metrics_v521["bleu"]), 0.65, avg(metrics_v521["bleu"]) - avg(metrics_v520["bleu"])])
    rs.append(["haeyo_ratio", avg(metrics_v520["haeyo_ratio"]), avg(metrics_v521["haeyo_ratio"]), 0.5, avg(metrics_v521["haeyo_ratio"]) - avg(metrics_v520["haeyo_ratio"])])
    rs.append(["phone_removal", avg(metrics_v520["phone_removal"]), avg(metrics_v521["phone_removal"]), 1.0, avg(metrics_v521["phone_removal"]) - avg(metrics_v520["phone_removal"])])
    rs.append(["formal_greeting", avg(metrics_v520["formal_greeting"]), avg(metrics_v521["formal_greeting"]), 0.05, avg(metrics_v521["formal_greeting"]) - avg(metrics_v520["formal_greeting"])])
    report_path = EVAL_DIR / "metrics_report.xlsx"
    report_wb.save(str(report_path))
    print(f"  보고서: {report_path}")
    print()
    print("VOC 4분기 정확도는 출력에 분기 라벨 포함된 경우만 측정. 별도 도구 필요.")


if __name__ == "__main__":
    main()
