"""
인사이트 ㉓ 재검증 — 정답데이터 빈도 측정으로 운영 룰 채택 여부 확인

검증 4 항목:
1. 시스템 피동 OK 7개 자의 확장 (처리/발송/이체/취소/접수/진행/정정)
   - 가이드 99p에 "체결되었습니다" 1개만 명시. 7개는 자의 확장.
   - 정답데이터 TOBE에 반복 등장하면 인사이트 ㉓로 채택 정당.

2. VOC 4분기 트리거 (~라고 알고/거듭/대단히/진심으로/금감원/언론/법적)
   - 정답데이터 ASIS에 반복 등장하면 트리거 룰 정당.

3. v5.20 시스템 능동 전환 룰 (처리됐 vs 처리했)
   - 정답데이터 TOBE 빈도로 v5.20 룰의 정답데이터 일치 여부 검증.

4. 톤 매트릭스 메시지 유형별 톤 차등 (해요체·하십시오체 분포)
   - 정답데이터 TOBE에서 메시지 유형별 톤 사용 분포 측정.
"""

from openpyxl import load_workbook
from pathlib import Path
import re

XLSX_PATHS = [
    Path(r"C:\dev\guide\정답데이터1\정답데이터1.xlsx"),
    Path(r"C:\dev\guide\정답데이터2\정답데이터2.xlsx"),
]

# ASIS·TOBE 컬럼 헤더 키워드
ASIS_KEYS = ["ASIS", "최종 메시지"]
TOBE_KEYS = ["TOBE", "개선안"]


def find_col(headers, keywords):
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        h_str = str(h).upper()
        if any(kw.upper() in h_str for kw in keywords):
            return i
    return None


def collect_text(xlsx_paths, col_keys):
    """모든 xlsx의 모든 시트에서 col_keys 컬럼 텍스트 모두 수집"""
    all_text = []
    for path in xlsx_paths:
        wb = load_workbook(str(path), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            col = find_col(headers, col_keys)
            if col is None:
                continue
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v:
                    all_text.append(str(v))
    return all_text


def count_pattern(texts, pattern):
    """리스트의 모든 텍스트에서 pattern (정규식) 매치 횟수"""
    if isinstance(pattern, str):
        return sum(t.count(pattern) for t in texts)
    return sum(len(pattern.findall(t)) for t in texts)


def main():
    print("=== 인사이트 ㉓ 재검증 시작 ===\n")

    print("[1/4] 정답데이터 ASIS·TOBE 텍스트 수집...")
    asis_texts = collect_text(XLSX_PATHS, ASIS_KEYS)
    tobe_texts = collect_text(XLSX_PATHS, TOBE_KEYS)
    print(f"  ASIS: {len(asis_texts)} cells / TOBE: {len(tobe_texts)} cells\n")

    # ===== 검증 1: 시스템 피동 OK 8개 =====
    print("=== 검증 1: 시스템 피동 OK 화이트리스트 빈도 ===")
    passive_words = [
        ("처리되었습니다", "처리했습니다"),
        ("발송되었습니다", "발송했습니다"),
        ("체결되었습니다", "체결했습니다"),  # 가이드 99p 명시
        ("이체되었습니다", "이체했습니다"),
        ("취소되었습니다", "취소했습니다"),
        ("접수되었습니다", "접수했습니다"),
        ("정정되었습니다", "정정했습니다"),
        ("진행되지 않았습니다", "진행하지 않았습니다"),
    ]
    print(f"  {'표현':<25} {'TOBE 피동':<12} {'TOBE 능동':<12} {'채택?':<8}")
    print(f"  {'-'*60}")
    for passive, active in passive_words:
        p_cnt = count_pattern(tobe_texts, passive)
        a_cnt = count_pattern(tobe_texts, active)
        verdict = "유지" if p_cnt > a_cnt else "검토"
        marker = " ★99p★" if "체결" in passive else ""
        print(f"  {passive:<25} {p_cnt:<12} {a_cnt:<12} {verdict}{marker}")

    # ===== 검증 2: VOC 4분기 트리거 =====
    print("\n=== 검증 2: VOC 4분기 트리거 키워드 빈도 (ASIS) ===")
    voc_triggers = [
        ("~라고 알고", "VOC_misclaim"),
        ("~인 줄 알", "VOC_misclaim"),
        ("~라고 해서", "VOC_misclaim"),
        ("거듭", "VOC_edit"),
        ("대단히", "VOC_edit"),
        ("진심으로", "VOC_edit"),
        ("금감원", "VOC_edit"),
        ("언론", "VOC_edit"),
        ("법적", "VOC_edit"),
    ]
    for trigger, scenario in voc_triggers:
        # ~ 처리
        kw = trigger.replace("~", "")
        cnt = count_pattern(asis_texts, kw)
        print(f"  '{trigger}' → {scenario}: ASIS {cnt}회")

    # ===== 검증 3: v5.20 시스템 능동 전환 룰 =====
    print("\n=== 검증 3: v5.20 시스템 능동 전환 룰 빈도 ===")
    active_rules = [
        ("발송되었습니다", "보내 드렸습니다"),
        ("안내드립니다", "알려드립니다"),
        ("안내드립니다", "안내합니다"),
        ("등록되었습니다", "등록했습니다"),
        ("등록되었습니다", "등록하셨습니다"),
    ]
    print(f"  {'원래 표현':<20} {'대안 표현':<20} {'TOBE 원래':<10} {'TOBE 대안':<10}")
    print(f"  {'-'*60}")
    for original, alternative in active_rules:
        o_cnt = count_pattern(tobe_texts, original)
        a_cnt = count_pattern(tobe_texts, alternative)
        print(f"  {original:<20} {alternative:<20} {o_cnt:<10} {a_cnt:<10}")

    # ===== 검증 4: 톤 매트릭스 (해요체·하십시오체) =====
    print("\n=== 검증 4: 톤 매트릭스 — 해요체·하십시오체 분포 ===")
    HAEYOCHE_PATTERNS = [r"해 주세요", r"확인해 주세요", r"신청해 주세요", r"이용해 주세요"]
    HASIPSIO_PATTERNS = [r"바랍니다", r"드립니다", r"하십시오"]

    asis_haeyo = sum(count_pattern(asis_texts, re.compile(p)) for p in HAEYOCHE_PATTERNS)
    asis_hasipsio = sum(count_pattern(asis_texts, re.compile(p)) for p in HASIPSIO_PATTERNS)
    tobe_haeyo = sum(count_pattern(tobe_texts, re.compile(p)) for p in HAEYOCHE_PATTERNS)
    tobe_hasipsio = sum(count_pattern(tobe_texts, re.compile(p)) for p in HASIPSIO_PATTERNS)

    print(f"  ASIS  해요체: {asis_haeyo:>6} / 하십시오체·격식체: {asis_hasipsio:>6}")
    print(f"  TOBE  해요체: {tobe_haeyo:>6} / 하십시오체·격식체: {tobe_hasipsio:>6}")
    if asis_haeyo + asis_hasipsio > 0:
        asis_ratio = asis_haeyo / (asis_haeyo + asis_hasipsio)
        print(f"  ASIS  해요체 비율: {asis_ratio:.1%}")
    if tobe_haeyo + tobe_hasipsio > 0:
        tobe_ratio = tobe_haeyo / (tobe_haeyo + tobe_hasipsio)
        print(f"  TOBE  해요체 비율: {tobe_ratio:.1%}")
    print(f"  → 톤 매트릭스 (행동 요청 시 해요체) 정답데이터 검증: {'적용됨 ✅' if tobe_haeyo > asis_haeyo * 5 else '검토 필요'}")

    print("\n=== 인사이트 ㉓ 재검증 완료 ===")
    print("\n결과 해석:")
    print("- 검증 1: TOBE 피동 > TOBE 능동이면 시스템 주체 피동 OK 화이트리스트 채택 정당")
    print("- 검증 2: ASIS에 트리거 키워드 등장 빈도로 VOC 4분기 분기 작동 가능성 확인")
    print("- 검증 3: TOBE 대안 > TOBE 원래면 v5.20 능동 전환 룰 정답데이터 일치")
    print("- 검증 4: TOBE 해요체 비율이 ASIS 대비 크게 증가하면 톤 매트릭스 적용 정당")


if __name__ == "__main__":
    main()
