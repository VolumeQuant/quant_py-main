"""
v5.21 평가셋 50건 추출 (메모리 Phase 7·8 5주차 결정)

기준:
- 정답데이터2.xlsx 모든 시트에서 협의결과 = "검토필요" 또는 blank 행 추출
- 도메인 분포 균형 50건 선정
- vector DB 미반영 (LLM이 새로 보는 케이스)

출력: output/평가셋/eval_50_cases.xlsx (msg_code·도메인·ASIS·TOBE·검토 사유 컬럼)
       output/평가셋/baseline_v5.20_outputs_template.xlsx (사용자가 v5.20·v5.21 출력 채울 자리)
"""

from openpyxl import load_workbook, Workbook
from pathlib import Path
import random

XLSX_PATH = Path(r"C:\dev\guide\정답데이터2\정답데이터2.xlsx")
OUT_DIR = Path(r"C:\dev\guide\output\평가셋")

# 시트별 도메인 매핑 (시트명 키워드 기반)
DOMAIN_MAP = {
    "1차_국내해외주식채권": "stocks_bonds",
    "2차_파생 등": "derivatives",
    "3차_프로세스 외": "process_misc",
    "4차_ 연금 전체": "pension",
}

# 컬럼 인덱스 (1차 시트 기준, 다른 시트는 다를 수 있음)
COL_MSG_CODE = "B"
COL_ASIS = "H"
COL_TOBE = "I"
COL_AGREE = "U"  # 협의결과 컬럼


def col_to_idx(col_letter):
    """A=1, B=2, ..."""
    return ord(col_letter) - ord("A") + 1


def get_domain(sheet_name):
    for kw, dom in DOMAIN_MAP.items():
        if kw in sheet_name:
            return dom
    return "other"


def extract_eval_candidates(xlsx_path):
    """검토필요·blank 행 추출"""
    wb = load_workbook(str(xlsx_path), data_only=True)
    candidates = []
    for sheet_name in wb.sheetnames:
        if not any(kw in sheet_name for kw in DOMAIN_MAP):
            continue
        ws = wb[sheet_name]
        domain = get_domain(sheet_name)

        # 헤더 확인 (1행)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # "협의결과" 컬럼 위치 찾기
        agree_col = None
        for i, h in enumerate(headers, start=1):
            if h and "협의결과" in str(h):
                agree_col = i
                break
        if agree_col is None:
            print(f"  WARN: {sheet_name} 협의결과 컬럼 없음, 스킵")
            continue

        # ASIS·TOBE·msg_code 컬럼 찾기
        asis_col = None
        tobe_col = None
        msg_col = None
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            h_str = str(h)
            if "ASIS" in h_str.upper() or "최종 메시지" in h_str:
                asis_col = i
            elif "TOBE" in h_str.upper() or "개선안" in h_str:
                tobe_col = i
            elif "메시지코드" in h_str:
                msg_col = i

        if not (asis_col and tobe_col):
            print(f"  WARN: {sheet_name} ASIS/TOBE 컬럼 없음, 스킵")
            continue

        # 데이터 추출
        for row in range(2, ws.max_row + 1):
            agree = ws.cell(row=row, column=agree_col).value
            asis = ws.cell(row=row, column=asis_col).value
            tobe = ws.cell(row=row, column=tobe_col).value
            msg = ws.cell(row=row, column=msg_col).value if msg_col else f"{sheet_name[:5]}_{row}"

            agree_str = str(agree).strip() if agree else ""
            # 검토필요 또는 blank 추출 (협의결과 비어있거나 "검토" 키워드 포함)
            if agree_str in ["", "검토필요"] or "검토" in agree_str:
                if asis and tobe:
                    candidates.append(
                        {
                            "domain": domain,
                            "sheet": sheet_name,
                            "msg_code": msg,
                            "asis": str(asis).strip(),
                            "tobe": str(tobe).strip(),
                            "agree_reason": agree_str,
                        }
                    )

    return candidates


def select_50(candidates):
    """도메인 분포 균형 50건 선정"""
    by_domain = {}
    for c in candidates:
        by_domain.setdefault(c["domain"], []).append(c)

    print(f"  후보 분포: {[(d, len(cs)) for d, cs in by_domain.items()]}")

    # 도메인별 N건 (50/도메인수 + 잉여)
    n_domains = len(by_domain)
    base = 50 // n_domains
    extra = 50 - base * n_domains

    selected = []
    random.seed(42)
    for i, (dom, cs) in enumerate(by_domain.items()):
        n = base + (1 if i < extra else 0)
        random.shuffle(cs)
        selected.extend(cs[:n])

    return selected[:50]


def save_xlsx(rows, output_path, columns):
    wb = Workbook()
    ws = wb.active
    ws.title = "eval_set"
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    wb.save(str(output_path))
    print(f"  OK: {output_path} ({len(rows)} rows)")


def main():
    print("=== 평가셋 50건 추출 ===\n")
    OUT_DIR.mkdir(exist_ok=True)

    candidates = extract_eval_candidates(XLSX_PATH)
    print(f"  검토필요·blank 후보 추출: {len(candidates)}건")

    if len(candidates) < 50:
        print(f"  WARN: 후보 {len(candidates)}건 < 50건 — 모두 사용")
        selected = candidates
    else:
        selected = select_50(candidates)

    # eval set 저장
    eval_path = OUT_DIR / "eval_50_cases.xlsx"
    save_xlsx(
        selected,
        eval_path,
        ["domain", "sheet", "msg_code", "asis", "tobe", "agree_reason"],
    )

    # baseline template 저장 (사용자가 v5.20·v5.21 출력 채울 자리)
    template_rows = [
        {
            "msg_code": r["msg_code"],
            "domain": r["domain"],
            "asis": r["asis"],
            "tobe_reference": r["tobe"],
            "v5.20_output": "[채워 넣기]",
            "v5.21_output": "[채워 넣기]",
        }
        for r in selected
    ]
    template_path = OUT_DIR / "baseline_v5.20_outputs_template.xlsx"
    save_xlsx(
        template_rows,
        template_path,
        ["msg_code", "domain", "asis", "tobe_reference", "v5.20_output", "v5.21_output"],
    )

    print(f"\n=== 완료 ===")
    print(f"  평가셋: {eval_path}")
    print(f"  baseline 템플릿: {template_path}")
    print(f"\n사용 방법:")
    print(f"  1. baseline_v5.20_outputs_template.xlsx 열어서 v5.20·v5.21 LLM 출력 채워 넣기")
    print(f"  2. measure_metrics.py 실행해 5 메트릭 측정")


if __name__ == "__main__":
    main()
